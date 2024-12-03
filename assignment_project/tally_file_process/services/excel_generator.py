import openpyxl
from openpyxl.utils import get_column_letter

## This class is responsible for generating the Excel file.
class ExcelGenerator:
    def __init__(self, data):
        self.data = data
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Vouchers"

    def generate(self):
        headers = self.data[0].keys()
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            self.sheet[f"{col_letter}1"] = header

        for row_num, record in enumerate(self.data, 2):
            for col_num, (key, value) in enumerate(record.items(), 1):
                col_letter = get_column_letter(col_num)
                self.sheet[f"{col_letter}{row_num}"] = value

    def save(self, file_path):
        self.workbook.save(file_path)
